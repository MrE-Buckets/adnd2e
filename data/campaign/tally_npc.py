import pandas as pd
import openpyxl
import os

def write_summary(df_main, gen_name, folder_path):
    """
    Stand-alone function to generate the Master Summary 
    using a template and a consolidated DataFrame.
    """
    summary_template_path = "data/_master_summary.xlsx"
    if os.path.exists(summary_template_path):
        wb = openpyxl.load_workbook(summary_template_path)
        ws_sum = wb.active
        ws_sum["A1"].value = gen_name
        
        # Use the corrected exact-match matrix
        current_matrix = calculate_internal_matrix(df_main)
        
        row_map = {"Commoner": 3, "Elite": 4, "Adventurer": 5, "Hero": 6}
        races = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]
        
        for p_key, target_row in row_map.items():
            for i, r_key in enumerate(races):
                count_col = 2 + (i * 2) # Matches the B, D, F, H, J, L structure
                ws_sum.cell(row=target_row, column=count_col).value = current_matrix[p_key][r_key]
        
        summary_name = os.path.join(folder_path, f"{gen_name}_Summary.xlsx")
        wb.save(summary_name)
        print(f"Master Summary generated at: {summary_name}")

def apply_npc_styling(writer, sheet_name, df):
    """Meticulous Formatting for the 10-column structure."""
    workbook = writer.book
    ws = writer.sheets[sheet_name]
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1, 'align': 'center'})
    ws.set_column('A:E', 12)
    ws.set_column('F:F', 16)
    ws.set_column('G:G', 22)
    ws.set_column('H:H', 10)
    ws.set_column('I:I', 12)
    ws.set_column('J:J', 11)
    for col_num, value in enumerate(df.columns):
        ws.write(0, col_num, value, header_fmt)
    ws.freeze_panes(1, 0)

def calculate_internal_matrix(df):
    """Calculates counts for the summary using strict exact matching."""
    matrix = {}
    races = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]
    tiers = ["Commoner", "Elite", "Adventurer", "Hero"]
    
    for tier in tiers:
        matrix[tier] = {}
        for race in races:
            # FIX: We use exact equality '==' and f-string to match the 'Race: ' format.
            # This prevents 'Half-Elf' from being counted as 'Elf'.
            target_race_string = f"{race}: "
            count = len(df[(df['Tier'] == tier) & (df['Race'] == target_race_string)])
            matrix[tier][race] = count
    return matrix

def export_npc_pop(lists, gen_name, matrix, folder_path=None):
    """Exports data with a Logic Gate to Omit Init/Prime at the Kingdom level."""
    headers = ['Race', 'Sex', 'Age', 'Weight', 'Height', 'Alignment', 'Stats', 'Tier lvl.', 'Tier', 'Death']
    raw_rows = lists[gen_name]
    df_main = pd.DataFrame(raw_rows, columns=headers)

    if not folder_path:
        folder_path = os.path.join("data/campaign", gen_name, "npc_population")
    os.makedirs(folder_path, exist_ok=True)

    # --- THE LOGIC GATE ---
    # Blocks Init and Prime only if we are in the kingdom-level 'npc_population' folder.
    is_kingdom_root = folder_path.endswith("npc_population")

    if not is_kingdom_root:
        for suffix in ["Init", "Prime"]:
            file_path = os.path.join(folder_path, f"{gen_name}_{suffix}.xlsx")
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                df_main.to_excel(writer, sheet_name=gen_name, index=False)
                apply_npc_styling(writer, gen_name, df_main)

    write_summary(df_main, gen_name, folder_path)
    # --- SUMMARY GENERATION ---
    # summary_template_path = "data/_master_summary.xlsx"
    # if os.path.exists(summary_template_path):
    #     wb = openpyxl.load_workbook(summary_template_path)
    #     ws_sum = wb.active
    #     ws_sum["A1"].value = gen_name
        
    #     # Use the corrected exact-match matrix
    #     current_matrix = calculate_internal_matrix(df_main)
        
    #     row_map = {"Commoner": 3, "Elite": 4, "Adventurer": 5, "Hero": 6}
    #     races = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]
        
    #     for p_key, target_row in row_map.items():
    #         for i, r_key in enumerate(races):
    #             count_col = 2 + (i * 2)
    #             ws_sum.cell(row=target_row, column=count_col).value = current_matrix[p_key][r_key]
        
    #     summary_name = os.path.join(folder_path, f"{gen_name}_Summary.xlsx")
    #     wb.save(summary_name)